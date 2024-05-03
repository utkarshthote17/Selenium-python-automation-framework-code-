import inspect
import logging
import softest
import csv
from openpyxl import load_workbook, Workbook


class Util(softest.TestCase):

    def asserListItems(self, listobject, valuetocheck):
        value = None
        for i in listobject:
            print("The Text is : " + i.text)
            if valuetocheck == 1:
                value = "1 Stop"
            elif valuetocheck == 2:
                value = "2 Stops"

            self.soft_assert(self.assertEqual, i.text, value)

            if i.text == value:
                print("Assert Passed")
            else:
                print("Assert Failed")

        self.assert_all()


    def Custom_logger():

        loggername = inspect.stack()[1][3]
        logger = logging.getLogger(loggername)
        logger.setLevel(logging.DEBUG)
        ch = logging.StreamHandler()
        fh = logging.FileHandler("demolog1.log")
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                                      datefmt='%A %d %b %Y %H %M %S %p')
        formatter1 = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

        ch.setFormatter(formatter)
        fh.setFormatter(formatter1)
        logger.addHandler(ch)
        logger.addHandler(fh)
        return logger


    def read_data_from_excel(file_name, sheet_name):
        datalist = []
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        rows = sheet.max_row
        cols = sheet.max_column
        print(rows)
        print(cols)
        for i in range(1, rows + 1):
            row = []
            for j in range(1, cols + 1):
                row.append(sheet.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist


    def read_data_from_csv(file):
        datalist = []
        with open(file,'r') as f:
            reader = csv.reader(f)
            for row in reader:
                datalist.append(row)
        return datalist

